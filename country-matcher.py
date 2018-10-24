import pycountry
from fuzzywuzzy import process
from openpyxl import load_workbook
import matplotlib.pyplot as plt; plt.rcdefaults()
import matplotlib.pyplot as plt
import pandas as pd
import sys
from datetime import datetime


# Find best match for country
def find_best_match(misspelled, correct_names):
    closest, ratio = process.extractOne(misspelled, correct_names)
    corrected = False
    if misspelled != closest:
        corrected = True

    # Manual corrections - adjusting output
    if "Tanzania" in closest:
        closest = "Tanzania"
        corrected = True

    if "Venezuela" in closest:
        closest = "Venezuela"
        corrected = True

    if "Palestine" in closest:
        closest = "Palestine"
        corrected = True

    if "Russian" in closest:
        closest = "Russia"
        corrected = True

    if "Sint Maarten" in closest:
        closest = "Sint Maarten"
        corrected = True

    if "Macedonia" in closest:
        closest = "Macedonia"
        corrected = True

    if "Bolivia" in closest:
        closest = "Bolivia"
        corrected = True

    if "Korea" in closest:
        closest = "Korea"
        corrected = True

    if "Taiwan" in closest:
        closest = "Taiwan"
        corrected = True

    if "Iran" in closest:
        closest = "Iran"
        corrected = True

    if "Congo" in closest:
        closest = "Congo"
        corrected = True

    if closest == "United States":
        closest = "USA"
        corrected = True

    if closest == "United Arab Emirates":
        closest = "UAE"
        corrected = True

    if closest == "United Kingdom":
        closest = "UK"
        corrected = True

    return closest, ratio, corrected

# Create a chart based on the extracted data
def generate_chart(data_to_graph, graph_save_name, graph_title):
    plt.figure(figsize=(16, 9))
    ax = data_to_graph.plot(kind="bar", width=0.9)
    current_date = str(datetime.now().strftime('%m/%d/%Y'))

    plt.ylabel('Signups', fontsize=14, labelpad=20)
    plt.title(graph_title + ' as of ' + current_date)

    rects = ax.patches

    for rect in rects:
        # Get X and Y placement of label from rect.
        y_value = rect.get_height()
        x_value = rect.get_x() + rect.get_width() / 2

        # Number of points between bar and label. Change to your liking.
        space = 5
        # Vertical alignment for positive values
        va = 'bottom'

        # If value of bar is negative: Place label below bar
        if y_value < 0:
            # Invert space to place label below
            space *= -1
            # Vertically align label at top
            va = 'top'

        # Create annotation
        plt.annotate(
            y_value,                    # Use `label` as label
            (x_value, y_value),         # Place label at end of the bar
            xytext=(0, space),          # Vertically shift label by `space`
            textcoords="offset points", # Interpret `xytext` as offset in points
            ha='center',                # Horizontally center label
            va=va)                      # Vertically align label differently for
                                        # positive and negative values.

    plt.savefig(graph_save_name + '.png', bbox_inches='tight', dpi=300)

# Read from the processed .csv files instead of re-processing the entire Excel document
run_local = 0

if run_local == 0:
    # Build a list of all the country names
    country_name_list = []
    for idx, val in enumerate(pycountry.countries):
        country_name_list.append(val.name)

    # Excel file to be processed
    excel_file = '/Users/derek/Downloads/OCT3_CFC_org_country.xlsx'
    wb = load_workbook(excel_file)

    # Sheet name
    sheet_name = wb['Sheet 1']

    # Total number of rows
    r = sheet_name.max_row

    # Lists to hold data after processing
    corrected_list_high_confidence = []
    corrected_list_low_confidence = []

    # Counter for output
    num = 1

    # Read each cell
    for i in range(3, r + 1):
        cell = str(sheet_name.cell(row=i, column=1).value)
        cell = cell.lower().strip()

        # Manual processing - adjusting output
        if cell == "usa" or cell == "america" or cell == "texas" or "usa" in cell:
            cell = "United States"

        if cell == "uae":
            cell = "United Arab Emirates"

        if cell == "italia":
            cell = "Italy"

        if cell == "vancouver":
            cell = "Canada"

        if cell ==  "lima" or "lima" in cell:
            cell = "Peru"

        if cell == "london" or cell == "england":
            cell = "United Kingdom"

        if cell == "hyderabad" or cell == "ineiq" or cell == "nellore" or cell == "kakinada":
            cell = "India"

        if cell == "ozzy ostrich":
            cell = "Australia"

        if cell == "chendu":
            cell = "China"

        if cell == "españa":
            cell = "Spain"

        results = find_best_match(cell, country_name_list)

        # results[0] - corrected country output
        # results[1] - confidence match from original country to corrected country
        # results[2] - bool to indicate if original country was modified

        # If confidence level is 80% or higher, match is likely good, add to the high confidence list
        # If confidence level is below 80%, original country data is likely bad, add to low confidence list
        if results[1] >= 80:
            corrected_list_high_confidence.append(results[0])
        else:
            corrected_list_low_confidence.append([cell, results[0]])

        # Print each item
        print(num, cell, " - ", results[0], " - ", results[1])

        # Increment the counter
        num += 1
        
    country_totals = pd.Series(corrected_list_high_confidence)
    country_totals_bad = pd.Series(corrected_list_low_confidence)
    print("")
    print("***********")
    print('Total registrations:', num - 1)
    print('Good data registrations:', len(corrected_list_high_confidence))
    print('Bad data registrations:',len(corrected_list_low_confidence))
    print("***********")
    print("")
    country_totals.to_csv('country_totals_all.csv')
    country_totals_bad.to_csv('country_totals_all_bad.csv')
else:
    country_totals = pd.Series.from_csv('country_totals_all.csv')
    country_totals_bad = pd.Series.from_csv('country_totals_all_bad.csv')
    print("")
    print("***********")
    print('Good data registrations:', len(country_totals))
    print('Bad data registrations:',len(country_totals_bad))
    print("***********")
    print("")

country_totals_all = country_totals.value_counts()
country_totals_all.to_csv('country_totals_all_values.csv')

country_totals_top_fourty = country_totals.value_counts().nlargest(40)

print(country_totals_all)

chart_name = 'country_totals_top_40'+ '_' + str(datetime.now().strftime('%b_%d_%y'))

# DEBUG
#print("***********")
#print("Creating charts...")
generate_chart(country_totals_top_fourty, chart_name, 'Top 40 CFC Signups by Country')

# Generate index.html file manually to show the extracted data and corresponding chart
stdout = sys.stdout
sys.stdout = open('index.html','w')
print('<!DOCTYPE html>\n'
'<html>\n'
'  <body>\n'
'  <style>\n'
'    section {\n'
'      width: 100%;\n'
'      margin: auto;\n'
'      padding: 10px;\n'
'    }\n'
'    div#one {\n'
'      width: 25%;\n'
'      float: left;\n'
'    }\n'
'    div#two {\n'
'      margin-left: 15%;\n'
'    }\n'
'  </style>\n'
'  <section>\n'
'    <div id="one">\n'
'      <table>\n'
'        <tr>\n'
'          <td><b>', 'Total Registrations:', (len(country_totals) + len(country_totals_bad)), '</b></td>\n'
'        </tr>\n')

# Print the country and number of registrations into the index.html file
counter = 1
for i, v in country_totals_all.items():
    print('        <tr>')
    print('          <td>', counter, i, '</td>')
    print('          <td>', v, '</td>')
    print('        </tr>')
    counter += 1
    
print('      </table>\n'
'    </div>')

print('      <div id="two">\n'
'        <img src="' + chart_name + '.png"' + ' style="width:1200px;height:750px;">\n'
'      </div>\n'
'    </section>\n'
'  </body>\n'
'</html>')

sys.stdout = stdout
print('')
print("***********")
print("Complete!")
